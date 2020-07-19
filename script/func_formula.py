# -*- coding: utf-8 -*-
"""
Created on Tue Jul  7 00:05:42 2020

@author: wenfan
"""


import pandas as pd
import os
import sys
import numpy as np
from openpyxl.utils import get_column_letter
 
from func_absolute_value import *

def get_processed_df_with_formula(df,
                                  dept_num):
        
    df_with_row = df.copy()
    
    df_with_row['ROW_NUMBER'] = df_with_row.index+2        

    try:
        df_header, start_index = get_clean_dataframe(df_with_row,
                                'date',
                                -1)
    except ValueError:
        df_header, _ = get_clean_dataframe(df_with_row,
                            'year',
                            -1)
    df_header.columns = list(df_header.columns)[:-1]+['ROW_NUMBER']
   
    columns_orig = [item.upper() if isinstance(item, str) else item for item in df_header.columns ]
    
    df_processed = preprocess_df(df_header,
                            -1,
                            ['ROW_NUMBER'])
            
    year_month_list = list(set([item for item in zip(df_processed.YEAR, df_processed.MONTH)]))

    output_columns = ['DEPARTMENT ID',
                        'YEAR',
                        'MONTH',
                        'PAYMENT',
                        'BUDGET']
    output_df = pd.DataFrame(columns =output_columns)
    
    for i, (year, month) in enumerate(year_month_list):
        
        output_df.loc[i,'DEPARTMENT ID'] = dept_num
        output_df.loc[i,'YEAR'] = year
        output_df.loc[i,'MONTH'] = month
        
        section = df_processed[((df_processed.YEAR==year) & (df_processed.MONTH==month))]
        rows = section['ROW_NUMBER']
        rows = sorted(rows)
        
        payment_header = [ item for item in columns_orig if isinstance(item, str) and (item.find('PAYMENT')>-1 or item.find('SPENDING')>-1) and item.find('CUM')==-1][0]        
        column_letter_payment = get_column_letter(columns_orig.index(payment_header)+1)
        output_df.loc[i,'PAYMENT'] = f"=SUM('Support Dept {dept_num}'!{column_letter_payment}{rows[0]}:{column_letter_payment}{rows[-1]})"
        #print(f'original header is {columns_orig}, \n payment column is {column_letter_payment}')

        budget_max = section['BUDGET'].max()
        budget_row = rows[list(section['BUDGET']).index(budget_max)]
        
        budget_header = [ item for item in columns_orig if isinstance(item, str) and item.find('BUDGET')>-1 and item.find('CUM')==-1][0]
        column_letter_budget = get_column_letter(columns_orig.index(budget_header)+1)
        output_df.loc[i,'BUDGET'] = F"='Support Dept {dept_num}'!{column_letter_budget}{budget_row}"
        #print(f'original header is {columns_orig}, \n budfet column is {column_letter_budget}')
            
    return output_df


def add_cum_rolling_columns(report_output):
    
    report_output.reset_index(inplace = True, drop=True)    
    columns = list(report_output.columns)
    report_output['ROW NUMBER'] = report_output.index+2
    report_output['DIFFERENCE'] = report_output['ROW NUMBER'].apply(lambda x: f"={get_column_letter(columns.index('BUDGET')+1)}{x}"\
                                                                                 +f"-{get_column_letter(columns.index('PAYMENT')+1)}{x}")

    for index, row in report_output.iterrows():
        dept_id = row['DEPARTMENT ID']
        year = row['YEAR']
        current_row = row['ROW NUMBER']
        rows_per_id = list(report_output.loc[report_output['DEPARTMENT ID']==dept_id]['ROW NUMBER'])
        rows_per_id = sorted(rows_per_id)

        rows_per_year = list(report_output.loc[(report_output['DEPARTMENT ID']==dept_id)&(report_output['YEAR']==year)]['ROW NUMBER'])
        rows_per_year = sorted(rows_per_year)

        column_payment = get_column_letter(columns.index('PAYMENT')+1)
        report_output.loc[index, 'CUMULATIVE PAYMENT'] = f"=SUM({column_payment}{rows_per_year[0]}:{column_payment}{current_row})"
        
        column_budget = get_column_letter(columns.index('BUDGET')+1)
        report_output.loc[index, 'CUMULATIVE BUDGET'] = f"=SUM({column_budget}{rows_per_year[0]}:{column_budget}{current_row})"

        column_cum_payment = get_column_letter(columns.index('PAYMENT')+1+3)
        column_cum_budget = get_column_letter(columns.index('BUDGET')+1+3)        
        report_output.loc[index, 'CUMULATIVE DIFFERENCE'] = f"={column_cum_budget}{current_row}-{column_cum_payment}{current_row}"
        
        if current_row-2<rows_per_id[0]:
            report_output.loc[index, 'PAYMENT ROLLING 3 PERIODS'] = 0
            report_output.loc[index, 'BUDGET ROLLING 3 PERIODS'] = 0
        
        else:
            report_output.loc[index, 'PAYMENT ROLLING 3 PERIODS'] = f"=SUM({column_payment}{current_row-2}:{column_payment}{current_row})"
            report_output.loc[index, 'BUDGET ROLLING 3 PERIODS'] = f"=SUM({column_budget}{current_row-2}:{column_budget}{current_row})"
        
    report_output = report_output.drop(columns=['ROW NUMBER'])
    
    return report_output
